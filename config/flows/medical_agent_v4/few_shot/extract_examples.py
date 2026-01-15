"""
从提示词文件中提取示例并生成Excel文档
"""
import re
import os
from pathlib import Path
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 项目根目录
PROJECT_ROOT = Path(__file__).parent.parent.parent.parent.parent
PROMPTS_DIR = PROJECT_ROOT / "config" / "flows" / "medical_agent_v3" / "prompts"
OUTPUT_DIR = Path(__file__).parent
OUTPUT_FILE = OUTPUT_DIR / "提示词案例库.xlsx"


def extract_intent_examples(content: str) -> List[Dict[str, Any]]:
    """从意图识别提示词中提取示例"""
    examples = []
    
    # 提取各意图类型的示例
    intent_patterns = [
        (r'## 1\. record.*?\n\n\*\*示例\*\*：(.*?)(?=\n\n##|\n\n#|$)', 'record'),
        (r'## 2\. query.*?\n\n\*\*示例\*\*：(.*?)(?=\n\n##|\n\n#|$)', 'query'),
        (r'## 3\. qa.*?\n\n\*\*示例\*\*：(.*?)(?=\n\n##|\n\n#|$)', 'qa'),
        (r'## 4\. greeting.*?\n\n\*\*示例\*\*：(.*?)(?=\n\n##|\n\n#|$)', 'greeting'),
    ]
    
    for pattern, intent in intent_patterns:
        match = re.search(pattern, content, re.DOTALL)
        if match:
            example_text = match.group(1).strip()
            # 提取每个示例（以"- "开头）
            example_lines = re.findall(r'-\s*"([^"]+)"', example_text)
            for line in example_lines:
                examples.append({
                    '用户输入': line,
                    '识别意图': intent,
                    '置信度': '高',
                    '关键词匹配': '是',
                    '备注': ''
                })
    
    return examples


def extract_record_examples(content: str) -> List[Dict[str, Any]]:
    """从记录Agent提示词中提取示例"""
    examples = []
    
    # 提取使用场景示例
    # 记录药品示例
    medication_pattern = r'\*\*使用场景示例\*\*：\s*\n- "([^"]+)" → ([^\n]+)'
    medication_matches = re.findall(medication_pattern, content)
    for user_input, params in medication_matches:
        examples.append({
            '用户输入': user_input,
            'Agent回复': f'已记录：{params}',
            '记录类型': 'record_medication',
            '数据完整性': '完整',
            '备注': ''
        })
    
    # 记录症状示例
    symptom_pattern = r'## 4\. 记录症状.*?\*\*使用场景示例\*\*：\s*\n- "([^"]+)" → ([^\n]+)'
    symptom_matches = re.findall(symptom_pattern, content, re.DOTALL)
    for user_input, params in symptom_matches:
        examples.append({
            '用户输入': user_input,
            'Agent回复': f'已记录：{params}',
            '记录类型': 'record_symptom',
            '数据完整性': '完整',
            '备注': ''
        })
    
    # 记录健康事件示例
    event_pattern = r'## 5\. 记录健康事件.*?\*\*使用场景示例\*\*：\s*\n- "([^"]+)" → ([^\n]+)'
    event_matches = re.findall(event_pattern, content, re.DOTALL)
    for user_input, params in event_matches:
        examples.append({
            '用户输入': user_input,
            'Agent回复': f'已记录：{params}',
            '记录类型': 'record_health_event',
            '数据完整性': '完整',
            '备注': ''
        })
    
    # 提取输出示例（JSON格式）
    json_examples = re.findall(r'\*\*场景\d+：([^*]+)\*\*\s*```json\s*({[^}]+})\s*```', content, re.DOTALL)
    for scenario, json_content in json_examples:
        # 解析JSON内容
        response_match = re.search(r'"response_content":\s*"([^"]+)"', json_content)
        reasoning_match = re.search(r'"reasoning_summary":\s*"([^"]+)"', json_content)
        record_success_match = re.search(r'"record_success":\s*(true|false)', json_content)
        record_type_match = re.search(r'"record_type":\s*"([^"]+)"', json_content)
        
        examples.append({
            '用户输入': scenario.strip(),
            'Agent回复': response_match.group(1) if response_match else '',
            '记录类型': record_type_match.group(1) if record_type_match else '',
            '数据完整性': '成功' if record_success_match and record_success_match.group(1) == 'true' else '需要询问',
            '备注': reasoning_match.group(1) if reasoning_match else ''
        })
    
    return examples


def extract_query_examples(content: str) -> List[Dict[str, Any]]:
    """从查询Agent提示词中提取示例"""
    examples = []
    
    # 提取各查询类型的示例
    query_patterns = [
        (r'## 1\. 查询血压记录.*?\*\*使用场景示例\*\*：\s*\n(.*?)(?=\n\n##|\n\n#|$)', 'query_blood_pressure'),
        (r'## 2\. 查询药品记录.*?\*\*使用场景示例\*\*：\s*\n(.*?)(?=\n\n##|\n\n#|$)', 'query_medication'),
        (r'## 3\. 查询症状记录.*?\*\*使用场景示例\*\*：\s*\n(.*?)(?=\n\n##|\n\n#|$)', 'query_symptom'),
        (r'## 4\. 查询健康事件记录.*?\*\*使用场景示例\*\*：\s*\n(.*?)(?=\n\n##|\n\n#|$)', 'query_health_event'),
    ]
    
    for pattern, query_type in query_patterns:
        match = re.search(pattern, content, re.DOTALL)
        if match:
            example_text = match.group(1).strip()
            # 提取每个示例行（格式：- "..." → ...）
            example_lines = re.findall(r'-\s*"([^"]+)"\s*→\s*([^\n]+)', example_text)
            for user_input, params in example_lines:
                examples.append({
                    '用户输入': user_input.strip(),
                    'Agent回复': f'已为您查询{query_type.replace("query_", "")}数据',
                    '查询类型': query_type,
                    '时间范围示例': params.strip(),
                    '备注': ''
                })
    
    return examples


def extract_qa_examples(content: str) -> List[Dict[str, Any]]:
    """从QA Agent提示词中提取示例"""
    examples = []
    
    # 提取场景示例和回复话术
    # 场景1：疾病诊断、治疗方案等
    scene1_pattern = r'\*\*场景1：([^*]+)\*\*\s*- 患者问题示例：\s*(.*?)\s*- \*\*回复话术\*\*[：]?\s*(.*?)(?=\n\n\*\*场景|\n\n###|\n\n#|$)'
    scene1_matches = re.findall(scene1_pattern, content, re.DOTALL)
    for scene_name, questions, reply in scene1_matches:
        # 提取问题示例（格式：- 疾病诊断，例如：...）
        question_lines = re.findall(r'-\s*([^，,]+)[，,]?\s*例如[：:]?\s*([^\n]+)', questions)
        if not question_lines:
            # 如果没有"例如"，直接提取问题描述
            question_lines = re.findall(r'-\s*([^\n]+)', questions)
            question_lines = [(q, '') for q in question_lines if '例如' not in q and len(q.strip()) > 5]
        
        for question_desc, question_example in question_lines:
            question_text = question_example.strip() if question_example else question_desc.strip()
            if question_text and len(question_text) > 3:
                examples.append({
                    '问题': question_text,
                    '答案': reply.strip()[:500],
                    '场景类型': '诊疗相关',
                    '关键词': question_desc.strip(),
                    '安全边界分类': '疾病诊断/治疗方案',
                    '备注': scene_name.strip()
                })
    
    # 提取其他场景（场景2-场景7等）
    other_scenes_pattern = r'\*\*场景\d+：([^*]+)\*\*\s*- 患者问题示例[：]?\s*(.*?)\s*- \*\*回复话术[：]?\*\*[：]?\s*(.*?)(?=\n\n\*\*场景|\n\n###|\n\n#|$)'
    other_matches = re.findall(other_scenes_pattern, content, re.DOTALL)
    for scene_name, questions, reply in other_matches:
        # 提取问题示例
        question_lines = re.findall(r'-\s*([^\n]+)', questions)
        for question in question_lines:
            question = question.strip()
            if question and len(question) > 5 and '例如' not in question:
                examples.append({
                    '问题': question,
                    '答案': reply.strip()[:500],
                    '场景类型': '药物相关',
                    '关键词': '',
                    '安全边界分类': scene_name.strip(),
                    '备注': ''
                })
    
    # 提取紧急情况示例（简化处理）
    emergency_sections = re.findall(r'#### (一、[^#]+)', content, re.DOTALL)
    for section in emergency_sections:
        # 提取症状名称和回复话术
        symptom_items = re.findall(r'\*\*(\d+)\.\s+([^*]+)\*\*\s*-.*?\s*- \*\*回复话术\*\*[：]?\s*(.*?)(?=\n\n\*\*|\n\n####|\n\n#|$)', section, re.DOTALL)
        for num, symptom, reply in symptom_items:
            examples.append({
                '问题': f'出现{symptom.strip()}症状',
                '答案': reply.strip()[:500],
                '场景类型': '危重症等紧急情况',
                '关键词': symptom.strip(),
                '安全边界分类': '紧急情况',
                '备注': ''
            })
    
    return examples


def extract_safety_boundary_examples(content: str) -> List[Dict[str, Any]]:
    """从QA Agent提示词中提取安全边界场景示例"""
    examples = []
    
    # 提取场景描述和回复话术
    scene_pattern = r'\*\*场景\d+：([^*]+)\*\*\s*-.*?\s*- \*\*回复话术[：]?\*\*[：]?\s*(.*?)(?=\n\n\*\*场景|\n\n###|\n\n#|$)'
    matches = re.findall(scene_pattern, content, re.DOTALL)
    
    for scene_desc, reply in matches:
        # 提取触发条件
        condition_match = re.search(r'触发条件[：:]?\s*(.*?)(?=\n|\*\*)', scene_desc, re.DOTALL)
        trigger_condition = condition_match.group(1).strip() if condition_match else ''
        
        examples.append({
            '场景描述': scene_desc[:200],
            '触发条件': trigger_condition[:200],
            '回复话术': reply.strip()[:500],
            '风险等级': '高' if '紧急' in scene_desc or '危重症' in scene_desc else '中',
            '备注': ''
        })
    
    return examples


def extract_unclear_examples(content: str) -> List[Dict[str, Any]]:
    """从不明确意图Agent提示词中提取示例"""
    examples = []
    
    # 提取通用打招呼示例
    greeting_pattern = r'\*\*问题维度[：:]?\s*([^*]+)\*\*\s*-.*?回复话术[：:]?\s*(.*?)(?=\n\n\*\*问题维度|\n\n###|\n\n#|$)'
    greeting_matches = re.findall(greeting_pattern, content, re.DOTALL)
    
    for dimension, reply in greeting_matches:
        # 提取示例对话
        dialogue_pattern = r'-\s*P:\s*([^\n]+)\s*\n\s*-\s*D:\s*([^\n]+)'
        dialogues = re.findall(dialogue_pattern, reply)
        
        for user_input, agent_reply in dialogues:
            examples.append({
                '用户输入': user_input.strip(),
                '识别意图': 'greeting',
                '置信度': '中',
                '关键词匹配': '是',
                '备注': dimension.strip()
            })
    
    # 提取意图澄清示例
    clarify_pattern = r'\*\*场景\d+[：:]?\s*([^*]+)\*\*\s*- 用户："([^"]+)"\s*\n- 反问："([^"]+)"'
    clarify_matches = re.findall(clarify_pattern, content, re.DOTALL)
    
    for scenario, user_input, agent_reply in clarify_matches:
        examples.append({
            '用户输入': user_input.strip(),
            '识别意图': '需要澄清',
            '置信度': '0.5-0.8',
            '关键词匹配': '部分',
            '备注': scenario.strip()
        })
    
    return examples


def extract_after_record_examples(content: str) -> List[Dict[str, Any]]:
    """从记录后处理Agent提示词中提取示例"""
    examples = []
    
    # 提取血压点评示例
    bp_pattern = r'#### 血压场景,([^,]+),([^\n]+)\s*\n\*\*场景特征\*\*[：:]?\s*([^\n]+)\s*\n\*\*回复案例\*\*[：:]?\s*\n([^\n]+)'
    bp_matches = re.findall(bp_pattern, content, re.DOTALL)
    
    for category, subcategory, feature, reply in bp_matches:
        examples.append({
            '场景类型': category.strip(),
            '子类型': subcategory.strip(),
            '场景特征': feature.strip(),
            '回复案例': reply.strip(),
            '备注': ''
        })
    
    return examples


def create_excel_file():
    """创建Excel文件"""
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认sheet
    
    # 读取提示词文件
    intent_file = PROMPTS_DIR / "00-intent_recognition_agent.md"
    record_file = PROMPTS_DIR / "11-record_agent.md"
    query_file = PROMPTS_DIR / "20-query_agent.md"
    qa_file = PROMPTS_DIR / "50-QA_agent.md"
    unclear_file = PROMPTS_DIR / "90-unclear-agent.md"
    after_record_file = PROMPTS_DIR / "12-after_record_agent.md"
    
    # 提取示例
    all_data = {}
    
    if intent_file.exists():
        with open(intent_file, 'r', encoding='utf-8') as f:
            content = f.read()
            all_data['意图识别示例'] = extract_intent_examples(content)
    
    if record_file.exists():
        with open(record_file, 'r', encoding='utf-8') as f:
            content = f.read()
            all_data['记录示例'] = extract_record_examples(content)
    
    if query_file.exists():
        with open(query_file, 'r', encoding='utf-8') as f:
            content = f.read()
            all_data['查询示例'] = extract_query_examples(content)
    
    if qa_file.exists():
        with open(qa_file, 'r', encoding='utf-8') as f:
            content = f.read()
            all_data['QA示例'] = extract_qa_examples(content)
            all_data['安全边界示例'] = extract_safety_boundary_examples(content)
    
    if unclear_file.exists():
        with open(unclear_file, 'r', encoding='utf-8') as f:
            content = f.read()
            all_data['不明确意图示例'] = extract_unclear_examples(content)
    
    if after_record_file.exists():
        with open(after_record_file, 'r', encoding='utf-8') as f:
            content = f.read()
            all_data['记录后处理示例'] = extract_after_record_examples(content)
    
    # 创建各个Sheet
    sheet_configs = {
        '意图识别示例': ['用户输入', '识别意图', '置信度', '关键词匹配', '备注'],
        '记录示例': ['用户输入', 'Agent回复', '记录类型', '数据完整性', '备注'],
        '查询示例': ['用户输入', 'Agent回复', '查询类型', '时间范围示例', '备注'],
        'QA示例': ['问题', '答案', '场景类型', '关键词', '安全边界分类', '备注'],
        '安全边界示例': ['场景描述', '触发条件', '回复话术', '风险等级', '备注'],
        '不明确意图示例': ['用户输入', '识别意图', '置信度', '关键词匹配', '备注'],
        '记录后处理示例': ['场景类型', '子类型', '场景特征', '回复案例', '备注'],
    }
    
    for sheet_name, columns in sheet_configs.items():
        if sheet_name in all_data and all_data[sheet_name]:
            ws = wb.create_sheet(title=sheet_name)
            
            # 写入表头
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, col_name in enumerate(columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 写入数据
            for row_idx, row_data in enumerate(all_data[sheet_name], 2):
                for col_idx, col_name in enumerate(columns, 1):
                    value = row_data.get(col_name, '')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            
            # 调整列宽
            for col_idx in range(1, len(columns) + 1):
                col_letter = get_column_letter(col_idx)
                ws.column_dimensions[col_letter].width = 30
    
    # 保存文件
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"✅ Excel文件已生成: {OUTPUT_FILE}")
    print(f"\n各Sheet数据统计:")
    for sheet_name, data in all_data.items():
        print(f"  - {sheet_name}: {len(data)} 条")


if __name__ == "__main__":
    create_excel_file()
